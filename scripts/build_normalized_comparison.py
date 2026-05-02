#!/usr/bin/env python3
"""Normalize wandb comparison Excel + append additional wandb run summary.

Reads the existing comparison Excel (one sheet `comparison`), then:
  1. Drops rows whose metric is scale-incompatible (MAE, MSE, RMSE,
     Levenshtein, gen_loss).
  2. Rescales 0-100 metrics (BLEU/METEOR/ROUGE) to 0-1.
  3. Fetches wandb run summary (default: nf2psz5z) and appends one value
     column + three Δ columns.
  4. Applies 3-color conditional formatting (blue→white→red, diverging
     around 0 for Δ columns, 0→1 for value columns) so the whole workbook
     can be read as a heatmap.
  5. Writes a NEW Excel file (original preserved).
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
import wandb
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Metric classification
# ---------------------------------------------------------------------------
DROP_METRICS = {
    "levenshtein_score",  # 0–85, length-dependent
    "mae", "mse", "rmse", # regression, unit-dependent
    "gen_loss",           # unbounded loss
}

PCT_METRICS = {  # 0–100 → /100 to 0–1
    "bleu2", "bleu4", "bleu_selfies", "bleu_smiles",
    "meteor", "meteor_llada", "meteor_wordnet",
    "rouge1", "rouge2", "rougeL",
}

# Native 0–1 (kept as-is):
#   accuracy, f1, precision, recall, roc_auc, failure_rate,
#   validity_ratio, exact_match_ratio, MACCS_FTS, RDK_FTS, morgan_FTS


# ---------------------------------------------------------------------------
# wandb utilities
# ---------------------------------------------------------------------------
def parse_run_url(url: str) -> tuple[str, str, str]:
    parts = [p for p in urlparse(url).path.strip("/").split("/") if p]
    if len(parts) >= 4 and parts[-2] == "runs":
        return parts[-4], parts[-3], parts[-1]
    raise ValueError(f"could not parse wandb run url: {url}")


def fetch_summary(run_url: str) -> tuple[dict, str]:
    entity, project, run_id = parse_run_url(run_url)
    print(f"[wandb] fetching summary: {entity}/{project}/{run_id}")
    api = wandb.Api()
    run = api.run(f"{entity}/{project}/{run_id}")
    return dict(run.summary), run_id


def lookup_summary_value(summary: dict, dataset: str, metric: str, strategy: str):
    """Try multiple known key patterns and return the first match (or None)."""
    # Strategy column in Excel is {random, semi_ar, likelihood}; wandb uses
    # `low_confidence_{random|semi_ar}` as a prefix. `likelihood` strategy
    # is generally absent in these runs.
    if strategy in ("random", "semi_ar"):
        suffix = f"low_confidence_{strategy}"
    else:
        suffix = strategy  # likelihood — may not exist

    candidates = [
        # Pattern A: y0g03wux style — val/{dataset}/{prefix_strategy}/{metric}
        f"val/{dataset}/{suffix}/{metric}",
        # Pattern B: nf2psz5z style — val/{metric}/{dataset}/{prefix_strategy}
        f"val/{metric}/{dataset}/{suffix}",
        # Alternative variants (legacy)
        f"val/{dataset}/random_{strategy}/{metric}" if strategy in ("random","semi_ar") else None,
        f"val/{metric}/{dataset}/random_{strategy}" if strategy in ("random","semi_ar") else None,
    ]

    for k in candidates:
        if k is None:
            continue
        if k in summary:
            v = summary[k]
            if isinstance(v, (int, float)):
                return float(v), k
    return None, None


# ---------------------------------------------------------------------------
# Excel read/transform
# ---------------------------------------------------------------------------
# Original columns (row 1):
# A dataset | B metric | C strategy | D cur_e1 | E cur_e2 | F cur_e2(128step)
# G prior_e1 | H prior_e2 | I prior_final
# J Δ(cur_e2 - prior_e2) | K Δ(cur_e2 - prior_final)
# L Δ(cur_e2(128step)- prior_e2) | M Δ(cur_e2(128step)- prior_final)
VALUE_COLS_IN = ["cur_e1", "cur_e2", "cur_e2(128step)",
                 "prior_e1", "prior_e2", "prior_final"]
DELTA_COLS_IN = ["Δ(cur_e2 - prior_e2)", "Δ(cur_e2 - prior_final)",
                 "Δ(cur_e2(128step)- prior_e2)", "Δ(cur_e2(128step)- prior_final)"]


def read_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["comparison"]
    headers = [ws.cell(1, c).value for c in range(1, 14)]
    rows = []
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if not d:
            continue
        row = {h: ws.cell(r, i + 1).value for i, h in enumerate(headers)}
        rows.append(row)
    return rows


def transform(rows: list[dict], new_col_name: str, summary: dict) -> tuple[list[dict], dict]:
    """Filter + rescale + attach new column. Returns (rows, stats)."""
    out = []
    stats = {
        "total_in": len(rows),
        "dropped_metric": 0,
        "rescaled": 0,
        "new_col_matched": 0,
        "new_col_missing": 0,
        "unmatched_samples": [],
    }
    for row in rows:
        metric = row["metric"]
        if metric in DROP_METRICS:
            stats["dropped_metric"] += 1
            continue

        # rescale PCT metrics (/100) across all numeric columns
        if metric in PCT_METRICS:
            for col in VALUE_COLS_IN + DELTA_COLS_IN:
                v = row.get(col)
                if isinstance(v, (int, float)):
                    row[col] = v / 100.0
            stats["rescaled"] += 1

        # fetch new run's summary
        new_val, matched_key = lookup_summary_value(
            summary, row["dataset"], metric, row["strategy"]
        )
        if new_val is not None and metric in PCT_METRICS:
            new_val = new_val / 100.0
        row[new_col_name] = new_val

        if new_val is not None:
            stats["new_col_matched"] += 1
        else:
            stats["new_col_missing"] += 1
            if len(stats["unmatched_samples"]) < 10:
                stats["unmatched_samples"].append(
                    (row["dataset"], metric, row["strategy"])
                )

        # compute new Δ columns vs prior_e2, prior_final, cur_e2
        def _delta(a, b):
            if isinstance(a, (int, float)) and isinstance(b, (int, float)):
                return a - b
            return None

        row[f"Δ({new_col_name} - prior_e2)"]    = _delta(new_val, row.get("prior_e2"))
        row[f"Δ({new_col_name} - prior_final)"] = _delta(new_val, row.get("prior_final"))
        row[f"Δ({new_col_name} - cur_e2)"]      = _delta(new_val, row.get("cur_e2"))

        out.append(row)
    return out, stats


# ---------------------------------------------------------------------------
# Excel write + conditional formatting
# ---------------------------------------------------------------------------
# Diverging scale colors for Δ columns (blue min → white 0 → red max)
BLUE_DEEP = "1F4E78"
WHITE     = "FFFFFF"
RED_DEEP  = "C00000"
# Sequential scale for value columns (white 0 → red 1)
RED_LIGHT = "FCE4D6"


def write_excel(rows: list[dict], new_col_name: str, out_path: Path):
    delta_cols_new = [
        f"Δ({new_col_name} - prior_e2)",
        f"Δ({new_col_name} - prior_final)",
        f"Δ({new_col_name} - cur_e2)",
    ]
    value_cols_new = [new_col_name]

    all_cols = (
        ["dataset", "metric", "strategy"]
        + VALUE_COLS_IN + value_cols_new
        + DELTA_COLS_IN + delta_cols_new
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "comparison"

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for ci, col in enumerate(all_cols, start=1):
        c = ws.cell(1, ci, col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for ri, row in enumerate(rows, start=2):
        for ci, col in enumerate(all_cols, start=1):
            c = ws.cell(ri, ci, row.get(col))
            if col not in ("dataset", "metric", "strategy") and isinstance(row.get(col), (int, float)):
                c.number_format = "0.0000"

    # Column widths
    for ci, col in enumerate(all_cols, start=1):
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = max(12, min(26, len(col) + 2))

    ws.freeze_panes = "D2"

    # Conditional formatting
    n_rows = len(rows)
    last_row = n_rows + 1
    if n_rows > 0:
        # 3-color diverging scale for Δ columns (blue min, white 0, red max)
        diverge_rule = ColorScaleRule(
            start_type="min", start_color=BLUE_DEEP,
            mid_type="num", mid_value=0, mid_color=WHITE,
            end_type="max", end_color=RED_DEEP,
        )
        # 2-color sequential for value columns: white at 0 → red at 1
        value_rule = ColorScaleRule(
            start_type="num", start_value=0, start_color=WHITE,
            end_type="num", end_value=1, end_color=RED_LIGHT,
        )

        for col in DELTA_COLS_IN + delta_cols_new:
            if col in all_cols:
                idx = all_cols.index(col) + 1
                letter = get_column_letter(idx)
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{last_row}", diverge_rule
                )

        for col in VALUE_COLS_IN + value_cols_new:
            if col in all_cols:
                idx = all_cols.index(col) + 1
                letter = get_column_letter(idx)
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{last_row}", value_rule
                )

    wb.save(out_path)
    print(f"[save] {out_path}  ({n_rows} rows, {len(all_cols)} cols)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    default_input = (
        "/opt/EMNLP_MolDA/New_MolDA/results/"
        "wandb_MolDA2Epoch비교_comparison_y0g03wux_vs_krozx9yp_vs_uqywm810"
        "(e2_128s) (1).xlsx"
    )
    default_output = (
        "/opt/EMNLP_MolDA/New_MolDA/results/"
        "wandb_comparison_normalized_with_nf2psz5z.xlsx"
    )
    default_url = "https://wandb.ai/hj_ai/MolDA_pro6000/runs/nf2psz5z"

    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default=default_input)
    ap.add_argument("--output", default=default_output)
    ap.add_argument("--wandb-url", default=default_url)
    args = ap.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)
    assert in_path.exists(), f"input not found: {in_path}"

    print(f"[load] {in_path}")
    rows = read_rows(in_path)
    print(f"  loaded {len(rows)} rows")

    summary, run_id = fetch_summary(args.wandb_url)
    print(f"  summary keys: {len(summary)}")

    new_col = f"{run_id}_summary"
    out_rows, stats = transform(rows, new_col, summary)

    print("\n=== Transform Stats ===")
    print(f"  input rows      : {stats['total_in']}")
    print(f"  dropped (metric): {stats['dropped_metric']}")
    print(f"  kept            : {len(out_rows)}")
    print(f"  rescaled (PCT)  : {stats['rescaled']}")
    print(f"  new col matched : {stats['new_col_matched']}")
    print(f"  new col missing : {stats['new_col_missing']}")
    if stats["unmatched_samples"]:
        print("  unmatched examples (first 10):")
        for d, m, s in stats["unmatched_samples"]:
            print(f"    ({d!r}, {m!r}, {s!r})")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel(out_rows, new_col, out_path)


if __name__ == "__main__":
    main()
